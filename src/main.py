
import os
import time
import re
import random

from openpyxl import load_workbook, Workbook
from nltk import regexp_tokenize, word_tokenize
from os.path import exists

with open('data/doc1.txt', 'r') as f:
    doc1 = f.read()

with open('data/doc2.txt', 'r') as f:
    doc2 = f.read()

with open('data/doc3.txt', 'r') as f:
    doc3 = f.read()

with open('data/doc4.txt', 'r') as f:
    doc4 = f.read()

with open('data/doc5.txt', 'r') as f:
    doc5 = f.read()

with open('data/doc6.txt', 'r') as f:
    doc6 = f.read()

documents = [
    {"filename": 'doc1.txt', "text": doc1},
    {"filename": 'doc2.txt', "text": doc2},
    {"filename": 'doc3.txt', "text": doc3},
    {"filename": 'doc4.txt', "text": doc4},
    {"filename": 'doc5.txt', "text": doc5},
    {"filename": 'doc6.txt', "text": doc6}
]

all_docs = doc1 + ' ' + doc2 + ' ' + doc3 + ' ' + doc4 + ' ' + doc5 + ' ' + doc6

sentence_splits = regexp_tokenize(all_docs, pattern=r'\.(?:\s+|$)', gaps=True)


words = all_docs.split()


def removePunc(word):
    disallowed_chars = ' 1234567890.,;?!‘’“”"\'\xe2\x80\x9d\x99\x98\x9c'
    new_string = word
    for char in disallowed_chars:
        new_string = new_string.replace(char, '')

    new_string += ' '
    return new_string


# ? Interesting words in this case are long words without special characters
def letterCount(word):
    if len(word) < 7:
        return False
    else:
        return True


words_no_punc = list(map(removePunc, words))

words_filtered = list(filter(letterCount, words_no_punc))

unique_list = []


def find_unique(word):
    if word not in unique_list:
        unique_list.append(word)


for word in words_filtered:
    find_unique(word)

sorted_words = sorted(unique_list)


def wordCount(word):
    count = 0
    for i in words_filtered:
        if (word) in i:
            count += 1
    return {"word": word, "word_count": count}


word_counts = map(wordCount, sorted_words)

sorted_word_counts = sorted(
    word_counts, key=lambda d: d['word_count'], reverse=True)


def trimWords(word):
    if word['word_count'] > 19:
        return True
    else:
        return False


trimmed_sorted = list(filter(trimWords, sorted_word_counts))


def wordLocations(word):
    doc_list = []
    for doc in documents:
        if (word['word']) in doc['text']:
            doc_list.append(doc['filename'])
    return {
        "word": word['word'],
        "word_count": word['word_count'],
        "featured": ', '.join(doc_list)
    }


word_locations = list(map(wordLocations, trimmed_sorted))


def wordExample(word):
    sent_list = []
    final_list = []
    for sent in sentence_splits:
        if (word['word']) in sent:
            sent_list.append(sent)

    for i in range(3):
        rand_item = random.choice(sent_list)
        final_list.append(rand_item + '.')
        sent_list.remove(rand_item)
    return {
        "word": word['word'],
        "word_count": word['word_count'],
        "featured": word['featured'],
        "examples": ' \n\n'.join(final_list)
    }


word_examples = list(map(wordExample, word_locations))

final_data = word_examples
# ? Overwrite and Save data sheet


print('Checking if DataSheet exists... ' +
      str(exists('./data/DataSheet.xlsx')))

if(exists('./data/DataSheet.xlsx')):
    workbook = load_workbook(filename='./data/DataSheet.xlsx')
else:
    workbook = Workbook()


sheet = workbook['Sheet']

sheet['A1'] = 'Word'
sheet['B1'] = 'Count'
sheet['C1'] = 'Appears in'
sheet['D1'] = 'Examples'

for i in final_data:
    sheet['A' + str(final_data.index(i) + 2)] = i['word'].capitalize()
    sheet['B' + str(final_data.index(i) + 2)] = i['word_count']
    sheet['C' + str(final_data.index(i) + 2)] = i['featured']
    sheet['D' + str(final_data.index(i) + 2)] = i['examples']

print('Appended data...')


if(exists('./data/DataSheet.xlsx')):
    print('Removing file...')
    os.remove('./data/DataSheet.xlsx')
    time.sleep(2)
    print('Removed file')

workbook.save(filename='./data/DataSheet.xlsx')

print('Creating new xlsx file...')
print('Finished')
